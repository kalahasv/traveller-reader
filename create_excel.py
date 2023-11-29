import openpyxl as op
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side, PatternFill,Alignment
from openpyxl.utils import column_index_from_string
import re
ws = None
ap_name = 'HUNG NG.'
def create_excel(trav_df,p_df,due_date): #fill in the variable data
    global ws
    global ap_name
    workbook = op.load_workbook('TRAVELER TEMPLATE.xlsx')
    ws = workbook.active
    #res = str(df.loc[0,'Purchase Order'])

    #print( df.loc[0,'Purchase Order'])
    center = Alignment(horizontal='center', vertical='center')
    ws['G4'] = trav_df.loc[0,'Purchase Order']
    ws['G5'] = trav_df.loc[0,'Customer Part ID']
    ws['G6'] = format_pn(trav_df)
    ws['B9'] = trav_df.loc[0,'Quantity']
    ws['D13'].alignment = center
    ws['E13'].alignment = center
    ws['C9'] = trav_df.loc[0,'Due Date']
    ws['B13'] = trav_df.loc[0,'Material']
    ws.row_dimensions[13].height = 100
    ws.row_dimensions[13].adjust_height = True
    #ws['E13'] = calculate_quantity(trav_df)
    ws['D9'] = calculate_shop_due(due_date)

   

    #keeps track of current row being entered
    c_row = 14
    process_num = 1
    #//Create processes in excel
    for index in range(p_df.shape[0] - 1, -1, -1):
        p = p_df.at[index,'Process']
        d = p_df.at[index,'Description']
        if p_df.at[index,'Due Date'] != 'N/A':
            dd = p_df.at[index,'Due Date'].strftime('%m-%d-%Y')
        # Do something with the row
        match p:
            case 'Turning':
                c_row = create_sbox(process_num,c_row,p,d,dd)
                
            case 'Operations':
                c_row = create_op_box(process_num,c_row,dd)
            case 'Deburr':
                c_row = create_sbox(process_num,c_row,p,d,dd)
            case 'Finish':
                c_row = create_sbox(process_num,c_row,p,d,dd)
            case 'Inserts':
                c_row = create_sbox(process_num,c_row,p,d,dd)
            case 'Part Marking':
                c_row = create_sbox(process_num,c_row,p,d,dd)
            case 'Final Inspection':
                c_row = create_sbox(process_num,c_row,p,d,dd)
            case 'Bag and Tag':
                c_row = create_sbox(process_num,c_row,p,d,dd)
            case 'Notes':
                c_row = create_notes(c_row,d)
            case _:
                print('Error: Process without method')
                
        if index != 0:
            create_qc_check(c_row)
        c_row += 1
        process_num += 1

    c_row += 1
    #put in last line
    create_end_line(c_row+15,ap_name)
    
    
    p_id = trav_df.loc[0,'Customer Part ID']
    p_id = strip_non_alphanumerics(p_id)

    workbook.save(p_id + ".xlsx")





def strip_non_alphanumerics(input_string):
    return re.sub(r'[^a-zA-Z0-9]', '', input_string)

def create_notes(start_row,description):
    global ws
    cell = ws[ 'B'+ str(start_row)]
    cell.value = "Notes:"

    desc_cell = ws[ 'B'+ str(start_row+1)]
    desc_cell.value = description
    ws.row_dimensions[start_row+1].height = 100
    ws.row_dimensions[start_row+1].adjust_height = True
    ws.merge_cells('B' + str(start_row+1) + ':' + 'I' + str(start_row+1))

    return start_row+1
def create_end_line(start_row,op_name):
    global ws
    #merge cells
    ws.merge_cells('B' + str(start_row) + ':' + 'C' + str(start_row))

    subtitle_cols = {'B':'Traveler Rev A','D':'Approval:','E':op_name,'G':'PHD Machining, LLC.'}
    for col,value in subtitle_cols.items():
        m_cell = ws[col + str(start_row)]
        m_cell.value = value
        m_cell.font = op.styles.Font(bold=True)

def create_qc_check(start_row):
        global ws

        qc_cell = ws['B' + str(start_row)]
        qc_cell.font = op.styles.Font(bold=True,size = 16)
        qc_cell.value = "QC"
        
def create_sbox(index,start_row,process,desc,due_date): #creates a standard box for a process, returns end row
    global ws
    #create number
    end_row = start_row+3
    start_column = 'B'
    end_column = 'B'
    ws.merge_cells(start_column + str(start_row) + ':' + end_column + str(end_row))
    num_cell = ws[start_column + str(start_row)]
    num_cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')
    num_cell.font = op.styles.Font(bold=True)
    num_cell.font = op.styles.Font(size=20)
    num_cell.value = str(index)
   

    #create title
    t_start_col = 'C'
    t_end_col = 'E'
    ws.merge_cells(t_start_col + str(start_row) + ':' + t_end_col + str(start_row))
    t_cell = ws[t_start_col + str(start_row)]
    t_cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')
    t_cell.font = op.styles.Font(bold=True)
    t_cell.value = str(process)
    

    #create description
    d_start_col = 'C'
    d_end_col = 'E'
    d_start_row = start_row+1
    d_end_row = end_row
    ws.merge_cells(d_start_col + str(d_start_row) + ':' + d_end_col + str(d_end_row))
    d_cell =  ws[d_start_col + str(d_start_row)]
    d_cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')
    d_cell.value = str(desc)
   
    #Create Start Qty
    q_col = 'F'
    ws[q_col + str(start_row)] = 'Start Qty:'

    #Create End Qty
    eq_row = start_row+2
    ws[q_col + str(eq_row)] = 'End Qty:'
    

    #Create Operator
    op_start_col = 'G'
    op_end_col = 'H'
    op_start_row = start_row
    op_end_row = end_row
    ws[op_start_col + str(op_start_row)] = 'Operator'
    ws.merge_cells(op_start_col + str(op_start_row+1) + ':' + op_end_col + str(op_end_row))
    
    #Create Due Date
    dd_start_col = 'I'
    dd_end_col = 'I'
    dd_start_row = start_row
    dd_end_row = end_row
    ws[dd_start_col + str(dd_start_row)] = 'Due:'
    ws.merge_cells(dd_start_col + str(dd_start_row+1) + ':' + dd_end_col + str(dd_end_row))
    dd_cell =  ws[dd_start_col + str(dd_start_row+1)]
    dd_cell.value = str(due_date)
    dd_cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')

    #add border to the box
    add_border(start_row,end_row,'B','I')

    return end_row+1
def create_op_box(index,start_row,due_date):


    #create title
    end_row = start_row+14
    start_column = 'B'
    end_column = 'C'
    ws.merge_cells(start_column + str(start_row) + ':' + end_column + str(start_row+1))
    ws.merge_cells('D'+str(start_row) + ':' + 'I' + str(start_row+1))
    t_cell = ws[start_column + str(start_row)]
    t_cell.font = op.styles.Font(bold=True)
    t_cell.font = op.styles.Font(size=20)
    t_cell.value = "Operations:"

    

    #create number
    num_srow = start_row+2
    num_erow = end_row
    start_column = 'B'
    end_column = 'B'
    ws.merge_cells(start_column + str(num_srow) + ':' + end_column + str(num_erow))
    num_cell = ws[start_column + str(num_srow)]
    num_cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')
    num_cell.font = op.styles.Font(bold=True)
    num_cell.font = op.styles.Font(size=20)
    num_cell.value = str(index)

    #create machine complete
    m_srow = start_row+3
    start_column = 'C'
    m_cell = ws[start_column + str(m_srow)]
    m_cell.font = op.styles.Font(bold=True,size = 13)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
    m_cell.fill = yellow_fill
    m_cell.value = "Machine Complete"

    #create ops total
    m_srow = start_row+3
    start_column = 'D'
    m_cell = ws[start_column + str(m_srow)]
    m_cell.font = op.styles.Font(bold=True,size = 13)
    m_cell.value = "Total OPs:"

    #create due title
    m_srow = start_row+3
    start_column = 'F'
    m_cell = ws[start_column + str(m_srow)]
    m_cell.font = op.styles.Font(bold=True,size = 13)
    m_cell.value = "Due Date:"

    #create due date
    m_srow = start_row+3
    start_column = 'G'
    m_cell = ws[start_column + str(m_srow)]
    m_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red fill

    m_cell.value = due_date


    #create subtitles
    subtitle_cols = {'D':'Start Qty:','E':'End Qty:','F':'Setup By:','G':'Operator 1','H':'Operator 2','I':'Date finished'}
    s_srow = start_row+4
    for col,value in subtitle_cols.items():
        m_cell = ws[col + str(s_srow)]
        m_cell.value = value

    #create ops
    op_col = 'C'
    op_srow = start_row+5
    for i in range(0,10):
        m_cell = ws[op_col + str(op_srow+i)]
        m_cell.value = 'OP ' + str(i+1)

    #add border 
    add_border(start_row,end_row,'B','I')


    return end_row + 1    
def add_border(start_row,end_row,start_col,end_col):
    global ws

    border = Border(
    left=Side(border_style='thin', color='000000'),  # thin solid black line
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
    )

    for column_letter in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=column_letter).border = border

def format_pn(trav_df):
    part = trav_df.loc[0,'Part Name']
    part = part.split('.')
    return part[0]

def calculate_quantity(trav_df):
    quantity = trav_df.loc[0,'Quantity']
    if(quantity <= 10 ):
        shop_quantity = quantity 
    else:
        shop_quantity = quantity 

    return shop_quantity

def calculate_shop_due(due_date):
    shop_day = due_date - timedelta(days=1)
    shop_str = shop_day.strftime('%m/%d/%Y')
    shop_str = shop_str 
    return shop_str



