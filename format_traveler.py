#take job information and put it into the traveller form
'''
note: once job information is put into database, this will need 
to be updated to reflect that

'''
import openpyxl as op
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from style import get_header_font,greenFill

def create_excel(): #fill in the non variable data
    workbook = op.Workbook()
    ws = workbook.active
    img = Image('phd-logo.png')
    ws.add_image(img,'A1')

    ws['F2'].fill = greenFill()
    ws['F2'] = "Job Traveler"
    ws['F2'].font = get_header_font()
    ws.merge_cells('F2:H2')

    ws['B7'] = "Deliveries"



    ws['B8'] = 'Qty Due'

    workbook.save("Traveller-1.xlsx")
