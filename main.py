
from pathlib import Path
import numpy as np
import pandas as pd
from IPython.display import display
import tabula
import openpyxl as op
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta

trav_df = None
processes = [] #list of processes
p_df = None
due_date = None
timezone = None

def pdf_to_df(file_name):
    global trav_df
    df = tabula.read_pdf(file_name, pages='all')

    concatenated_df = df[0]
    for d in df[1:]:
        concatenated_df = pd.concat([concatenated_df,d], axis=1)
    #print(concatenated_df)
    trav_df = concatenated_df

def create_excel(): #fill in the variable data
    workbook = op.load_workbook('TRAVELER TEMPLATE.xlsx')
    ws = workbook.active
    #res = str(df.loc[0,'Purchase Order'])

    #print( df.loc[0,'Purchase Order'])
    ws['G4'] = trav_df.loc[0,'Purchase Order']
    ws['G5'] = trav_df.loc[0,'Customer Part ID']
    ws['G6'] = format_pn()
    ws['B9'] = trav_df.loc[0,'Quantity']
    ws['C9'] = trav_df.loc[0,'Due Date']
    ws['B13'] = trav_df.loc[0,'Material']
    ws['E13'] = calculate_quantity()
    ws['D9'] = calculate_shop_due()

    #first add processes to queue

    workbook.save("Traveller-1.xlsx")

def calculate_shop_due():
    shop_day = due_date - timedelta(days=1)
    shop_str = shop_day.strftime('%m/%d/%Y')
    shop_str = shop_str + "," + timezone
    return shop_str

def calculate_due_date():
    global due_date
    global timezone

    unformatted = trav_df.loc[0,'Due Date']
    date = unformatted.split(',')
    date_num = date[0]
   
    date_format = '%m/%d/%Y'
    due_date = datetime.strptime(date_num,date_format)
    timezone = date[1]
    #print("Date:",due_date)

def format_pn():
    part = trav_df.loc[0,'Part Name']
    part = part.split('.')
    return part[0]

def calculate_quantity():
    quantity = trav_df.loc[0,'Quantity']
    if(quantity <= 10 ):
        shop_quantity = quantity + 1
    else:
        shop_quantity = quantity + 2

    return shop_quantity

def find_deburr_description():
    desc = ""
    if ('Finish' in trav_df.columns and trav_df.loc[0,'Finish'] != 'Standard'):
        desc = 'Deburr and Clean'
    else:
        desc = 'Deburr'
    return desc

def pm_type():
    #print("Made it here")
    #print(trav_df.loc[0,'Part Marking'])
    if 'Engraving' in trav_df.loc[0,'Part Marking']:
        return 'EGR'
    elif 'Laser' in trav_df.loc[0,'Part Marking']:
        return 'LSR'
    
def create_queue(t_status): #create queue for all processes
    global processes

    if(t_status == True):
        processes.append("Turning")
    processes.append("Operations")
    processes.append("Deburr")

    #if part marking says engraving, append it here
    
    if('Part Marking' in trav_df.columns and pm_type() == 'EGR'):
        processes.append("Part Marking")
       
    if('Finish' in trav_df.columns and trav_df.loc[0,'Finish'] != 'Standard'):
        processes.append("Finish")

    if ('Inserts' in trav_df.columns):
        processes.append("Inserts")

    if ('Part Marking' in trav_df.columns and pm_type() == 'LSR'):
        processes.append("Part Marking")

    processes.append("Final Inspection")
    processes.append("Bag and Tag")

def create_p_df(line_items):
    global processes
    #dataframe with items in the format Process, Description, Due Date
    global p_df
    column_names = ['Process','Description','Due Date']
    p_df = pd.DataFrame(columns = column_names)
    
    #print("Processes:",processes)
    processes_r = processes[::-1] #going from last process to first process
    #print(trav_df.columns)
   
    for i in range(len(processes_r)):
        #print("Process",p)
        p = processes_r[i]
        match p:
            case 'Bag and Tag':
                #calculate date
                if line_items > 5:
                    dd = due_date - timedelta(days = 2)
                else:
                    dd = due_date - timedelta(days = 1)
                #description is always the same
                desc = "Package to prevent shipping damage"
            case 'Final Inspection':
                if line_items > 5:
                    dd = due_date - timedelta(days = 2)
                else:
                    dd = due_date - timedelta(days = 1)
                desc = "QC"
            case 'Inserts':
                day = p_df.loc[p_df['Process'] == 'Final Inspection', 'Due Date'].values[0]
                day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
                dd = day - timedelta(days = 2)

                desc = trav_df.loc[0,'Inserts']

                #print("Day", day)
            case 'Finish':
                #p = "Plating"
                desc = trav_df.loc[0,'Finish']
                #print out current dates, wait for input
                print('Finish:',desc)
                # Print all columns except= inserts
                if 'Inserts' in p_df['Process'].values:
                    result_df = p_df[p_df['Process'] != 'Inserts']
                    print('Result', result_df)
                else:
                    print(p_df)
                
                #note: will have to put input validation checking
                day_str = input("Please input plating due date in the format yyyy-mm-dd.")


                date_format = '%Y-%m-%d'
                dd = datetime.strptime(day_str,date_format)
                if 'Inserts' in p_df['Process'].values: #overwrite inserts due date with the plating date
                    p_df.loc[p_df['Process'] == 'Inserts','Due Date'] = dd
            
            case 'Part Marking':
                #it might include "Bag and Tag", but shouldn't include that ;note: this will have to manually reviewed since there's formatting
               
                desc = trav_df.loc[0,'Part Marking']
                desc = desc.replace("Bag and Tag","",1)

                next_p = processes_r[i-1]
                #print("Next process:",next_p)
                dd = find_pm_day(next_p)
                
            case 'Deburr':
                desc = find_deburr_description()
                
                #check process after deburring
                next_p = processes_r[i-1]
                print("Next process:",next_p)
                dd = find_deburr_day(next_p)

            case 'Operations':
                dd = find_op_day()
                desc = '' #no description for this

            case 'Turning':
                dd = find_turning_day()
                desc = 'Quan'
         
            case _:
                dd = due_date
                desc = "Not Done Yet"

        p_df['Due Date'] = pd.to_datetime(p_df['Due Date'])
        row = {"Process": p,"Description": desc,"Due Date":dd}
        #p_df = p_df.append(row, ignore_index=True)
        p_df.loc[len(p_df)] = row
        
    print("Process data frame:\n",p_df)

def find_deburr_day(next_p):
    match next_p:

        case 'Finish':

            #Deburr is 1 day before plating.
            day = p_df.loc[p_df['Process'] == 'Finish', 'Due Date'].values[0]
            day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
            dd = day - timedelta(days = 1)
            
        case 'Part Marking':
            #Deburr should be the same day as Part Marking
            day = p_df.loc[p_df['Process'] == 'Part Marking', 'Due Date'].values[0]
            day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
            dd = day
        
        case 'Inserts':
            #Deburr is same day as Inserts
            day = p_df.loc[p_df['Process'] == 'Inserts', 'Due Date'].values[0]
            day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
            dd = day
        
        case 'Final Inspection':
            day = p_df.loc[p_df['Process'] == 'Final Inspection', 'Due Date'].values[0]
            day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
            dd = day - timedelta(days = 2)
    
    return dd

def find_pm_day(next_p):
    #subtract one day from whatever is the next process
    day = p_df.loc[p_df['Process'] == next_p, 'Due Date'].values[0]
    day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
    dd = day - timedelta(days = 1)
    return dd

def find_op_day():
    day = p_df.loc[p_df['Process'] == 'Deburr', 'Due Date'].values[0]
    day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
    dd = day - timedelta(days = 1)
    return dd

def find_turning_day():
    day = p_df.loc[p_df['Process'] == 'Operations', 'Due Date'].values[0]
    day = datetime.utcfromtimestamp(day.astype('datetime64[s]').astype(int))
    dd = day - timedelta(days = 1)
    return dd
     
if __name__ == '__main__':

    file_name = '0571C6E-traveler.pdf' #OP - Deburr - Finish - Inserts - PM (Laser) - Final Inspection - Bag&Tag
    #file_name = '052BD56-traveler.pdf' #OP - Deburr - PM(Engraving) - Finish - Final Inspection - Bag&Tag
    #file_name = '057531C-traveler.pdf' # OP - Deburr - Final Inspection - Bag&Tag
    #file_name = '05695AD-traveler.pdf' # OP - Deburr - Finish - Final - Bag&Tag
   
    
    path =  Path('jfiles',file_name)
    pdf_to_df(path)
    calculate_due_date()
   
    print("JOB ID:",trav_df.loc[0,'Purchase Order'])
    turning_status = input("Turning? Y/N: ")
    if turning_status == 'Y':
        turning_status = True
    else:
        turning_status = False
    #DEBUG: print out the job id
    #print(trav_df.columns)
    create_queue(turning_status)
    line_items = input("Number of line items:") #note: add error checking to this
    #print final due date
    print("Final Due Date: ",due_date)
    create_p_df(int(line_items))
    create_excel()